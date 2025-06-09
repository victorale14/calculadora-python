'''

22 a biblioteca openpyxl serve para ler e escrever excel 2010 em arquivos xlsx/xlsm,
segundo o próprio site (https://openpyxl.readthedocs.io/en/stable/index.html)
workbook é o que seria a planilha excel, portanto, a funcionalidade load_workbook se trata de carregar a planilha 

(esse código funciona presumindo que o usuário leu anteriormente a planilha)

24 - 25 imporando do módulo tkinter a função "askdirectory"
27 cria a variável pasta com a função que pergunta o dierótio para procurar o arquivo
31 adiciona a extensão ".xlsx" junto do nome do arquivo com uma f-string, junto da pasta
32 wb é declarado como a função que carrega o arquivo, para fins de simplificação
33 a planilha é definida como o ativo de wb, ou seja, a a planilha que está sendo carregada está sendo definida como ativa
34 - 39 são printados os avisos e o input do texto que será mudado é pego, assim como o texto que será posto no local do antigo
40 define a variável que diz se o texto foi ou não alterado
41 - 42 checa a fileira e as 'células' nas fileiras, chegando por coluna na direção horizontal, quando acaba passa para a próxima linha na direção vertical
45 - 47 caso o valor presente na célula seja igual ao texto que se quer mudar, ele muda o texto para o definido anteriormente
48 - 49 a planilha é salva e alterado é definido para verdadeiro
50 - 51 se alterado for verdadeiro, avisa que foi um sucesso, se for falso, avisa que que o termo não foi achado.
53 e 54 caso não consiga por causa de erro em localizar o arquivo, avisa que não foi possível encontrar o arquivo e o código para.

'''
from openpyxl import load_workbook
from tkinter.filedialog import askdirectory

def editar():
    pasta = askdirectory(title='Selecione a pasta em que o arquivo excel está.')
    print('Qual o nome do arquivo que você gostaria de editar?')
    nome_arquivo = input('Digite o nome do arquivo: ')
    try:
        arquivo = f'{pasta}\{nome_arquivo}.xlsx'
        wb = load_workbook(arquivo)
        planilha = wb.active
        print(f'O diretório {arquivo} foi carregado com sucesso!')
        print(f'Arquivo {nome_arquivo} carregado com sucesso!')
        print('Qual texto gostaria de mudar?')
        texto = input('> ').lower()
        print(f'Digite o que gostaria de pôr no lugar de {texto}')
        mudar = input('> ').lower()
        alterado = False
        for fileira in planilha:
            for celula in fileira:
                valor = celula.value
                valor = str(valor).lower()
                if valor == texto:
                    mudar = mudar.title()
                    celula.value = mudar
                    wb.save(arquivo)
                    alterado = True
        if alterado: print('O documento foi alterado e salvo com sucesso! Recomendo que feche e abra o arquivo.')
        else: print('Não foi possível alterar o documento. O termo a ser mudado não foi encontrado, tente novamente.')

    except FileNotFoundError:
        print(f'O arquivo {nome_arquivo} não foi encontrado. Verifique o nome e tente novamente.')
